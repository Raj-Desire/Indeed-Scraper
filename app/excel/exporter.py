"""
Excel Exporter
==============
Generates a clean, professionally formatted Excel workbook (.xlsx)
containing all scraped job leads.
"""

from datetime import datetime
from pathlib import Path
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.models.job import JobPosting
from app.utils.logger import logger


class ExcelExporter:
    """Generates styled Excel workbooks for scraped job postings."""

    def export(self, jobs: list[JobPosting], output_dir: str = "outputs") -> Path:
        """
        Export job postings to an Excel workbook.

        Args:
            jobs: List of JobPosting objects.
            output_dir: Target directory path.

        Returns:
            Path object pointing to the generated .xlsx file.
        """
        output_path = Path(output_dir)
        output_path.mkdir(parents=True, exist_ok=True)

        date_str = datetime.now().strftime("%Y-%m-%d")
        file_path = output_path / f"Indeed_Job_Leads_{date_str}.xlsx"

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Job Leads"

        # Headers
        headers = [
            "Job Title",
            "Company",
            "Location/Remote Type",
            "Salary Range",
            "Industry",
            "Company Size",
            "Posted Date",
            "Job URL",
        ]

        # Header formatting
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align

        ws.row_dimensions[1].height = 28

        # Data rows
        row_font = Font(name="Calibri", size=10)
        link_font = Font(name="Calibri", size=10, color="0563C1", underline="single")
        border = Border(
            left=Side(style="thin", color="D3D3D3"),
            right=Side(style="thin", color="D3D3D3"),
            top=Side(style="thin", color="D3D3D3"),
            bottom=Side(style="thin", color="D3D3D3"),
        )

        for row_idx, job in enumerate(jobs, start=2):
            posted = job.posted_date.strftime("%Y-%m-%d") if job.posted_date else job.posted_date_raw or "Not listed"

            row_values = [
                job.job_title,
                job.company,
                job.location_remote_type,
                job.salary_range,
                job.industry,
                job.company_size,
                posted,
                job.job_url,
            ]

            for col_idx, val in enumerate(row_values, start=1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = border
                cell.font = row_font

                if col_idx == 8 and str(val).startswith("http"):  # Job URL hyperlink
                    cell.value = "View on Indeed"
                    cell.hyperlink = str(val)
                    cell.font = link_font
                else:
                    cell.value = val

        # Auto-adjust column widths
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val = str(cell.value or "")
                if len(val) > max_len:
                    max_len = len(val)
            ws.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 50)

        # Freeze header row
        ws.freeze_panes = "A2"

        wb.save(file_path)
        logger.info("Excel exported: {} ({} leads)", file_path, len(jobs))
        return file_path
